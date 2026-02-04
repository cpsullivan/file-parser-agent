"""Excel spreadsheet parser using openpyxl."""

from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from .base_parser import BaseParser


class ExcelParser(BaseParser):
    """
    Parser for Excel spreadsheets (.xlsx, .xls).
    
    Extracts:
    - Data from all sheets
    - Sheet names and dimensions
    - Cell values with type preservation
    """
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    FILE_TYPE = 'excel'
    
    def parse(self) -> Dict[str, Any]:
        """
        Parse Excel file and extract content from all sheets.
        
        Returns:
            Dictionary containing:
            - filename, file_type, parsed_at (base fields)
            - metadata: Sheet names and count
            - sheets: List of sheet objects with name, dimensions, data
        """
        output = self.get_base_output()
        
        try:
            # Load workbook with data_only=True to get calculated values
            wb = load_workbook(self.filepath, data_only=True)
            
            # Extract metadata
            output['metadata'] = {
                'total_sheets': len(wb.sheetnames),
                'sheet_names': wb.sheetnames
            }
            
            # Extract data from each sheet
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = self._extract_sheet(ws, sheet_name)
                sheets.append(sheet_data)
            
            output['sheets'] = sheets
            
            wb.close()
            
        except Exception as e:
            output['error'] = f"Excel parsing error: {str(e)}"
            output['metadata'] = {'total_sheets': 0, 'sheet_names': []}
            output['sheets'] = []
        
        return output
    
    def _extract_sheet(self, ws: Worksheet, sheet_name: str) -> Dict[str, Any]:
        """
        Extract data from a single worksheet.
        
        Args:
            ws: openpyxl Worksheet instance
            sheet_name: Name of the worksheet
            
        Returns:
            Dictionary with name, dimensions, and data array
        """
        # Get dimensions
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        
        # Extract data
        data = []
        for row in ws.iter_rows(values_only=True):
            # Convert row to list, handling None and various types
            row_data = [self._convert_cell_value(cell) for cell in row]
            data.append(row_data)
        
        # Remove trailing empty rows
        while data and all(cell == '' for cell in data[-1]):
            data.pop()
        
        return {
            'name': sheet_name,
            'rows': len(data),
            'columns': max_col,
            'data': data
        }
    
    def _convert_cell_value(self, value: Any) -> Any:
        """
        Convert cell value to appropriate JSON-serializable type.
        
        Args:
            value: Cell value from openpyxl
            
        Returns:
            Converted value (str, int, float, bool, or '')
        """
        if value is None:
            return ''
        
        # Keep numbers and booleans as-is
        if isinstance(value, (int, float, bool)):
            return value
        
        # Convert datetime to ISO string
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        
        # Everything else becomes a string
        return str(value)
